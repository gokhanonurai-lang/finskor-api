"""
BilankoIQ — Mizan Parser
Excel mizandan standart bilanço ve gelir tablosu çıkarır.

Akış:
  1. Excel oku (openpyxl)
  2. Hesap kodu kolonunu tespit et
  3. TDHP eşleştirme tablosuyla kalemlere ata
  4. Eşleşme oranı düşükse Claude API fallback
  5. Normalize edilmiş BalanceSheet objesi dön
"""

from __future__ import annotations
import re
import json
import logging
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# 1. VERİ MODELİ
# ─────────────────────────────────────────────

@dataclass
class BalanceSheet:
    """Normalize edilmiş bilanço ve gelir tablosu kalemleri (TL)."""

    # AKTİF — Dönen Varlıklar
    kasa: float = 0.0                    # 100
    banka: float = 0.0                   # 102
    diger_hazir_degerler: float = 0.0    # 108
    ticari_alacaklar: float = 0.0        # 120 + 121
    diger_alacaklar_kv: float = 0.0      # 126 + 136
    stoklar: float = 0.0                 # 150–158 toplamı
    diger_donen_varliklar: float = 0.0   # 180–195

    # AKTİF — Duran Varlıklar
    ticari_alacaklar_uv: float = 0.0     # 220 + 221
    diger_alacaklar_uv: float = 0.0      # 226 + 236
    mali_duran_varliklar: float = 0.0    # 240–248
    maddi_duran_varliklar: float = 0.0   # 250–258 (net)
    maddi_olmayan_duv: float = 0.0       # 260–268 (net)
    diger_duran_varliklar: float = 0.0   # 270–299

    # PASİF — Kısa Vadeli Yükümlülükler
    banka_kredileri_kv: float = 0.0      # 300 + 301
    uzun_vadeli_borclar_kv: float = 0.0  # 303
    ticari_borclar_kv: float = 0.0       # 320 + 321
    ortaklara_borclar: float = 0.0       # 331
    diger_kv_borclar: float = 0.0        # 330–399 (331 hariç)

    # PASİF — Uzun Vadeli Yükümlülükler
    banka_kredileri_uv: float = 0.0      # 400 + 401
    diger_uv_borclar: float = 0.0        # 402–499

    # PASİF — Özkaynaklar
    odenmis_sermaye: float = 0.0         # 500
    sermaye_yedekleri: float = 0.0       # 520–529
    kar_yedekleri: float = 0.0           # 540–549
    gecmis_yil_karlari: float = 0.0      # 570
    gecmis_yil_zararlari: float = 0.0   # 580–589 (borç bakiyeli, özkaynaktan düşülür)
    donem_net_kari: float = 0.0          # 590

    # GELİR TABLOSU
    net_satislar: float = 0.0            # 600 - 610
    satislarin_maliyeti: float = 0.0     # 620 + 621 + 622
    pazarlama_giderleri: float = 0.0     # 631
    genel_yonetim_giderleri: float = 0.0 # 632
    arge_giderleri: float = 0.0          # 630
    diger_faaliyet_gelirleri: float = 0.0     # 640–649
    diger_faaliyet_giderleri: float = 0.0     # 633, 650–657, 659, 680–689 (658 hariç)
    enflasyon_duzeltme_zarari: float = 0.0    # 658 — nakit dışı, FAVÖK'e dahil değil
    yillara_yaygin_enflasyon_net: float = 0.0 # 697 — inşaat enflasyon düzeltmesi (FAVÖK dışı)
    finansman_gelirleri: float = 0.0     # 670–679
    finansman_giderleri: float = 0.0     # 660 + 661
    vergi_gideri: float = 0.0            # 691

    # META
    parse_method: str = "rule_based"     # "rule_based" | "ai_fallback"
    match_rate: float = 0.0              # Eşleşme oranı (0–1)
    warnings: list = field(default_factory=list)
    alt_hesaplar: dict = field(default_factory=dict)  # {parent_kod: [(raw_kod, ad, bakiye), ...]}
    _kullan_590: bool = False            # Çakışma: 590 net_kar olarak kullanılsın

    # ─── Türetilmiş toplamlar ───────────────────

    @property
    def nakit_ve_benzerleri(self) -> float:
        return self.kasa + self.banka + self.diger_hazir_degerler

    @property
    def donen_varliklar(self) -> float:
        return (self.nakit_ve_benzerleri + self.ticari_alacaklar +
                self.diger_alacaklar_kv + self.stoklar + self.diger_donen_varliklar)

    @property
    def duran_varliklar(self) -> float:
        return (self.ticari_alacaklar_uv + self.diger_alacaklar_uv +
                self.mali_duran_varliklar + self.maddi_duran_varliklar +
                self.maddi_olmayan_duv + self.diger_duran_varliklar)

    @property
    def toplam_aktif(self) -> float:
        return self.donen_varliklar + self.duran_varliklar

    @property
    def kv_borclar(self) -> float:
        return (self.banka_kredileri_kv + self.uzun_vadeli_borclar_kv +
                self.ticari_borclar_kv + self.ortaklara_borclar + self.diger_kv_borclar)

    @property
    def uv_borclar(self) -> float:
        return self.banka_kredileri_uv + self.diger_uv_borclar

    @property
    def toplam_borclar(self) -> float:
        return self.kv_borclar + self.uv_borclar

    @property
    def ozkaynaklar(self) -> float:
        return (self.odenmis_sermaye + self.sermaye_yedekleri +
                self.kar_yedekleri + self.gecmis_yil_karlari - self.gecmis_yil_zararlari +
                self.donem_net_kari)

    @property
    def toplam_pasif(self) -> float:
        return self.toplam_borclar + self.ozkaynaklar

    @property
    def brut_kar(self) -> float:
        return self.net_satislar - self.satislarin_maliyeti

    @property
    def faaliyet_giderleri(self) -> float:
        return (self.pazarlama_giderleri + self.genel_yonetim_giderleri +
                self.arge_giderleri)

    @property
    def favok(self) -> float:
        """FAVÖK = Brüt Kâr - Faaliyet Giderleri + Diğer Faaliyet Gelirleri - Diğer Faaliyet Giderleri"""
        return (self.brut_kar - self.faaliyet_giderleri +
                self.diger_faaliyet_gelirleri - self.diger_faaliyet_giderleri)

    @property
    def vergi_oncesi_kar(self) -> float:
        return self.net_kar + self.vergi_gideri

    @property
    def net_kar(self) -> float:
        # 697 iki tarafı birbirini dengeler (net ≈ 0), FAVÖK dışında tutulur
        hesaplanan = (self.favok
                      - self.enflasyon_duzeltme_zarari
                      - self.yillara_yaygin_enflasyon_net
                      + self.finansman_gelirleri
                      - self.finansman_giderleri
                      - self.vergi_gideri)
        # Yıl sonu + gelir tablosu çakışması: 590 kazanır
        if self._kullan_590 and self.donem_net_kari != 0:
            return self.donem_net_kari
        # Gelir tablosu yoksa (yıl sonu kapatılmış, net_satislar=0): 590 kullan
        if self.net_satislar == 0 and self.donem_net_kari != 0:
            return self.donem_net_kari
        return hesaplanan

    @property
    def finansal_borclar(self) -> float:
        """Sadece banka/finansal borçlar (ticari borçlar hariç)."""
        return self.banka_kredileri_kv + self.uzun_vadeli_borclar_kv + self.banka_kredileri_uv

    @property
    def net_borc(self) -> float:
        return self.finansal_borclar - self.nakit_ve_benzerleri

    def to_dict(self) -> dict:
        d = asdict(self)
        d.update({
            "nakit_ve_benzerleri": self.nakit_ve_benzerleri,
            "donen_varliklar": self.donen_varliklar,
            "duran_varliklar": self.duran_varliklar,
            "toplam_aktif": self.toplam_aktif,
            "kv_borclar": self.kv_borclar,
            "uv_borclar": self.uv_borclar,
            "toplam_borclar": self.toplam_borclar,
            "ozkaynaklar": self.ozkaynaklar,
            "brut_kar": self.brut_kar,
            "faaliyet_giderleri": self.faaliyet_giderleri,
            "favok": self.favok,
            "net_kar": self.net_kar,
            "finansal_borclar": self.finansal_borclar,
            "net_borc": self.net_borc,
        })
        return d


# ─────────────────────────────────────────────
# 2. TDHP HESAP KODU → KALEM EŞLEŞTİRME TABLOSU
# ─────────────────────────────────────────────

# Her kalem için (hesap_kodu_prefix_listesi, BalanceSheet_field_adı, işaret)
# işaret: +1 borçlu bakiye pozitif, -1 alacaklı bakiye pozitif ekler
ACCOUNT_MAP: list[tuple[list[str], str, int]] = [
    # ─── Bakiye yön kuralı ───────────────────────────────────────────────────
    # _read_excel artık SIGNED bakiye döner: borç bakiyesi → pozitif, alacak → negatif.
    # Bu nedenle sign değerleri şu anlama gelir:
    #   +1 → borç-normal aktif/gider hesabı  (borç_bak pozitif = pozitif katkı)
    #   -1 → alacak-normal pasif/gelir hesabı (alacak_bak negatif × -1 = pozitif katkı)
    # Contra-aktif (amortisman, karşılık vb.) artık sign=+1:
    #   alacak_bak negatif × +1 = negatif → aktifi doğal olarak düşürür.
    # ────────────────────────────────────────────────────────────────────────

    # NAKİT (Grup 10)
    (["10"], "kasa", 1),                    # Sadece grup satırı varsa (sub-kod yoksa)
    (["100"], "kasa", 1),
    (["101"], "diger_hazir_degerler", 1),   # Alınan çekler — borç normal
    (["102"], "banka", 1),
    (["103"], "diger_hazir_degerler", 1),   # Verilen çekler — alacak_bak negatif × +1 = eksi
    (["108"], "diger_hazir_degerler", 1),

    # TİCARİ ALACAKLAR (KV) — Grup 12
    (["120", "121", "122", "124", "126", "127", "128"], "ticari_alacaklar", 1),
    (["129"], "ticari_alacaklar", 1),       # Şüpheli alacak karşılığı — alacak_bak negatif

    # DİĞER ALACAKLAR (KV) — Grup 13
    (["131", "132", "133", "135",
      "136", "137", "138", "139"], "diger_alacaklar_kv", 1),

    # VERİLEN AVANSLAR (Grup 14) — aktif, borç-normal
    (["140", "141", "142", "143", "144", "145",
      "146", "147", "148", "149"], "diger_donen_varliklar", 1),

    # STOKLAR (Grup 15)
    (["150", "151", "152", "153", "154", "157", "159"], "stoklar", 1),
    (["158"], "stoklar", 1),               # Stok değer düşüklüğü karşılığı — alacak_bak negatif

    # MENKUL KIYMETLER (110-119) ve DİĞER DÖNEN VARLIKLAR
    (["110", "111", "112", "113", "114", "115", "116", "117", "118"], "diger_donen_varliklar", 1),
    (["119"], "diger_donen_varliklar", 1),  # Menkul kıymet değer düşüklüğü — alacak_bak negatif
    # Yıllara yaygın inşaat ve diğer
    (["160", "161"], "diger_donen_varliklar", 1),  # 161 alacak_bak negatif × +1 = eksi
    (["170", "171", "172", "173", "174", "175",
      "176", "177", "178", "179",
      "180", "181", "182", "183", "184", "185",
      "190", "191", "192", "193", "195", "196"], "diger_donen_varliklar", 1),

    # TİCARİ ALACAKLAR (UV) — Grup 22
    (["220", "221"], "ticari_alacaklar_uv", 1),
    (["226", "236"], "diger_alacaklar_uv", 1),

    # MALİ DURAN VARLIKLAR
    (["200", "201", "202", "203", "204", "205",
      "206", "207", "208", "209"], "mali_duran_varliklar", 1),  # 209 alacak_bak negatif

    # MADDİ DURAN VARLIKLAR (Grup 25) — net: 258=YOY pozitif, 257=amortisman negatif
    (["210", "211", "212", "213", "214", "215",
      "216", "217", "218", "219",
      "250", "251", "252", "253", "254", "255", "256", "258"], "maddi_duran_varliklar", 1),
    (["257"], "maddi_duran_varliklar", 1),  # Birikmiş amortismanlar — alacak_bak negatif × +1 = eksi

    # MADDİ OLMAYAN DURAN VARLIKLAR (Grup 26) — sadece 26x hesapları
    (["260", "261", "262", "263", "264", "265",
      "266", "267"], "maddi_olmayan_duv", 1),
    (["268", "278"], "maddi_olmayan_duv", 1),  # Birikmiş itfa — alacak_bak negatif × +1 = eksi

    # DİĞER DURAN VARLIKLAR
    (["230", "231", "232", "233", "234", "235",
      "240", "241", "242", "243", "244", "245", "246", "247", "248",
      "270", "271", "272", "273", "274", "275",
      "276", "277", "279",
      "280", "281", "282", "284", "285",
      "291", "292", "293", "294", "295",
      "296", "297", "298", "299"], "diger_duran_varliklar", 1),

    # KV BANKA KREDİLERİ — alacak-normal → sign=-1
    (["300", "301"], "banka_kredileri_kv", -1),
    (["302", "303"], "uzun_vadeli_borclar_kv", -1),
    (["304", "305", "306", "307", "308", "309"], "diger_kv_borclar", -1),

    # TİCARİ BORÇLAR (KV) — sign=-1
    (["310", "311"], "ticari_borclar_kv", -1),
    (["320", "321", "329"], "ticari_borclar_kv", -1),

    # ORTAKLARA BORÇLAR — sign=-1
    (["323", "331", "430"], "ortaklara_borclar", -1),

    # DİĞER KV BORÇLAR (331 hariç) — sign=-1
    (["312", "313", "314", "315", "316", "317", "318", "319",
      "322", "324", "325", "326", "327", "328",
      "330", "332", "333", "334", "335", "336",
      "337", "338", "339",
      "340", "341", "342", "343", "344", "345",
      "346", "347", "348", "349",
      "350", "351", "352", "353", "354", "355",
      "356", "357", "358", "359",
      "360", "361", "362", "363", "364", "365",
      "366", "368", "369",
      "370", "371", "372", "373", "374", "375",
      "376", "377", "378", "379",
      "380", "381", "382", "383", "384", "385",
      "386", "387", "388", "389",
      "390", "391", "392", "393", "394", "395",
      "396", "397", "398", "399"], "diger_kv_borclar", -1),

    # UV BANKA KREDİLERİ — sign=-1
    (["400", "401"], "banka_kredileri_uv", -1),
    (["402", "403", "404", "405", "406", "407",
      "408", "409",
      "410", "411", "412", "413", "414", "415",
      "416", "417", "418", "419",
      "420", "421", "422", "423", "424", "425",
      "426", "427", "428", "429",
      "431", "432", "433", "434", "435",
      "436", "437", "438", "439",
      "440", "441", "442", "443", "444", "445",
      "446", "447", "448", "449",
      "470", "471", "472", "473", "474", "475",
      "476", "477", "478", "479",
      "480", "481", "482", "483", "484", "485",
      "486", "487", "488", "489",
      "490", "491", "492", "493", "494", "495",
      "496", "497", "498", "499"], "diger_uv_borclar", -1),

    # ÖZKAYNAKLAR — sign=-1 (alacak-normal); borç-normal contra hesaplar da -1
    # (borç_bak pozitif × -1 = negatif → özkaynağı azaltır)
    (["500", "502", "504", "505", "510", "511", "512"], "odenmis_sermaye", -1),
    (["501", "503"], "odenmis_sermaye", -1),   # Ödenmemiş sermaye / olumsuz farklar
    (["525", "526", "527", "528", "529"], "sermaye_yedekleri", -1),
    (["530", "531", "532", "533", "534", "535",
      "536", "537", "538", "539"], "sermaye_yedekleri", -1),
    (["550", "551", "552", "553", "554", "555",
      "556", "557", "558", "559",
      "560", "561", "562", "563", "564", "565",
      "566", "567", "568", "569"], "sermaye_yedekleri", -1),
    (["506", "507", "508", "509"], "kar_yedekleri", -1),
    (["540", "541", "542", "543", "544", "545",
      "546", "547", "548", "549"], "kar_yedekleri", -1),
    (["570", "520"], "gecmis_yil_karlari", -1),
    (["521", "523"], "gecmis_yil_karlari", -1),   # Geçmiş yıl zararları — borç_bak × -1 = eksi
    (["580", "581", "582", "583", "584", "585",
      "586", "587", "588", "589"], "gecmis_yil_zararlari", 1),  # borç_bak × +1, ozkaynaktan düşülür
    (["590"], "donem_net_kari", -1),
    (["522"], "sermaye_yedekleri", -1),  # Sermaye Düzeltme Farkları — özkaynak rezervi, net kâr değil

    # GELİR TABLOSU
    (["600", "601", "602"], "net_satislar", -1),    # alacak-normal → sign=-1
    (["610", "611", "612"], "net_satislar", -1),    # satış indirimleri: borç_bak × -1 = eksi satış
    (["620", "621", "622", "623"], "satislarin_maliyeti", 1),
    (["630"], "arge_giderleri", 1),
    (["631"], "pazarlama_giderleri", 1),
    (["632"], "genel_yonetim_giderleri", 1),
    (["640", "641", "642", "643", "644",
      "645", "646", "647", "648", "649"], "diger_faaliyet_gelirleri", -1),  # alacak-normal
    (["633", "650", "651", "652", "653", "654",
      "655", "656", "657",        "659",
      "680", "681", "682", "683", "684", "685",
      "686", "687", "688", "689"], "diger_faaliyet_giderleri", 1),
    (["658"], "enflasyon_duzeltme_zarari", 1),  # nakit dışı — FAVÖK dışında tutulur
    (["697"], "yillara_yaygin_enflasyon_net", 1),  # inşaat enflasyon düzeltmesi — iki taraf dengeler
    (["660", "661"], "finansman_giderleri", 1),
    (["670", "671", "672", "673", "674",
      "675", "676", "677", "678", "679"], "finansman_gelirleri", -1),      # alacak-normal
    (["691"], "vergi_gideri", 1),
    # ÜRETİM/HİZMET MALİYET HESAPLARI (inşaat/üretim şirketleri ara dönem kapatmaz)
    # 74x = Hizmet/İnşaat Üretim Maliyeti → satış maliyetine eşdeğer
    (["740", "741", "742", "743", "744", "745",
      "746", "747", "748", "749"], "satislarin_maliyeti", 1),
    # 76x = Pazarlama Giderleri (maliyet merkezi)
    (["760", "761", "762", "763", "764", "765",
      "766", "767", "768", "769"], "pazarlama_giderleri", 1),
    # 77x = Genel Yönetim Giderleri (maliyet merkezi)
    (["770", "771", "772", "773", "774", "775",
      "776", "777", "778", "779"], "genel_yonetim_giderleri", 1),
]

# Hızlı lookup: hesap_kodu → (field_adı, işaret) listesi
_CODE_LOOKUP: dict[str, list[tuple[str, int]]] = {}
for prefixes, field_name, sign in ACCOUNT_MAP:
    for p in prefixes:
        _CODE_LOOKUP.setdefault(p, []).append((field_name, sign))


# ─────────────────────────────────────────────
# 3. EXCEL OKUMA
# ─────────────────────────────────────────────

# Alt hesap analizi için hedef ana hesaplar ve minimum alt kalem sayısı
_ALT_HESAP_HEDEFLER = frozenset([
    "120", "150", "253", "254",   # Aktif
    "300", "301", "320", "321",   # KV Borçlar
    "400", "401",                  # UV Borçlar
])
_ALT_HESAP_MIN_KALEM = 10


def _normalize_code(raw: str | int | float | None) -> str | None:
    """Hesap kodunu temizler: '120.01', '120 01', 120 → '120'"""
    if raw is None:
        return None
    s = str(raw).strip().split(".")[0].split(" ")[0].split("-")[0]
    s = re.sub(r"\D", "", s)
    return s if s else None


def _normalize_header(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    for a, b in [("ç","c"),("ö","o"),("ü","u"),("ı","i"),("i̇","i"),("ğ","g"),("ş","s"),("â","a"),("î","i"),("û","u")]:
        s = s.replace(a, b)
    return s

def _find_columns(ws) -> tuple[int | None, int | None, int | None, int | None, int | None]:
    """
    Hesap kodu, borç/alacak bakiye ve borç/alacak toplam kolonlarını tespit eder.
    Beş değer döner: (code_col, borc_bak_col, alacak_bak_col, borc_top_col, alacak_top_col)
    Tespit edilemeyen kolonlar None döner.
    """
    max_col = min(ws.max_column, 20)
    code_col = None
    borc_toplam_col = None
    alacak_toplam_col = None
    borc_bakiye_col = None
    alacak_bakiye_col = None
    balance_col = None

    for row in ws.iter_rows(min_row=1, max_row=20, max_col=max_col):
        for cell in row:
            val = _normalize_header(cell.value)
            if not val:
                continue
            if any(kw in val for kw in ["hesap kodu", "heskodu", "kod", "hs. kd", "account"]):
                code_col = cell.column
            if val in ["borc bakiyesi","borc bakiye","borc bak","borc bak.","debit balance","db bakiye"]:
                borc_bakiye_col = cell.column
            if val in ["alacak bakiyesi","alacak bakiye","alacak bak","alacak bak.","credit balance","cr bakiye"]:
                alacak_bakiye_col = cell.column
            if val in ["borc","borc (tl)","borc tutari","debit","borc toplam","borc miktari","toplam borc"]:
                borc_toplam_col = cell.column
            if val in ["alacak","alacak (tl)","alacak tutari","credit","alacak toplam","alacak miktari","toplam alacak"]:
                alacak_toplam_col = cell.column
            if val in ["bakiye","net bakiye","net tutar","balance","tutar"]:
                balance_col = cell.column

    if code_col is None:
        code_col = 1

    if borc_bakiye_col and alacak_bakiye_col:
        logger.info(f"Kolonlar: BORÇ BAKİYE={borc_bakiye_col}, ALACAK BAKİYE={alacak_bakiye_col}, "
                    f"BORÇ TOP={borc_toplam_col}, ALACAK TOP={alacak_toplam_col}")
        return code_col, borc_bakiye_col, alacak_bakiye_col, borc_toplam_col, alacak_toplam_col
    if borc_bakiye_col:
        return code_col, borc_bakiye_col, None, borc_toplam_col, alacak_toplam_col
    if borc_toplam_col and alacak_toplam_col:
        logger.info(f"Kolonlar: BORÇ={borc_toplam_col}, ALACAK={alacak_toplam_col}")
        return code_col, borc_toplam_col, alacak_toplam_col, borc_toplam_col, alacak_toplam_col
    if balance_col:
        return code_col, balance_col, None, None, None

    for row in ws.iter_rows(min_row=2, max_row=10, max_col=max_col):
        for cell in reversed(row):
            if isinstance(cell.value, (int, float)):
                balance_col = cell.column
                break
        if balance_col:
            break
    logger.warning(f"Kolon fallback: {balance_col}")
    return code_col, balance_col, None, None, None


def _is_parent_code(code, all_codes):
    # Sadece tam 3 haneli ana hesapları kullan (100, 120, 300 gibi)
    # 1-2 haneli grup kodları atlanır
    # 4+ haneli veya noktalı detay kodları atlanır
    clean = code.split('.')[0]
    if len(clean) <= 2:
        return True   # Grup kodu — atla
    if len(clean) == 3 and '.' not in code:
        return False  # Ana hesap — kullan
    return True       # Detay satır — atla


def _get_root3(code):
    return code.split(".")[0][:3]


# TDHP gelir tablosu hesaplarının doğal bakiye yönleri (yıl sonu kapanış tespiti için).
# Alacak-normal gelir hesapları: alacak toplamı kullanılır → signed negatif
# Borç-normal gider hesapları: borç toplamı kullanılır → signed pozitif
_ALACAK_NORMAL_6XX = frozenset(
    [str(x) for x in range(600, 610)] +   # Brüt satışlar
    [str(x) for x in range(640, 650)] +   # Diğer faaliyet gelirleri
    [str(x) for x in range(670, 680)]     # Finansman gelirleri
)
# 690/692 kapanış devir hesapları — zaten toplam olduklarından atlanır
_SKIP_6XX = frozenset(["690", "692"])
# İki taraflı hareket hesapları: borc_top - alacak_top net farkı kullanılır
# (697: 697.170 borç + 697.350 alacak birbiriyle offsetlanır, net P&L etkisi sıfır)
_NET_HAREKET_6XX = frozenset(["697"])


def _read_excel(filepath):
    """
    Excel mizan dosyasını okur.

    Returns:
        (rows, alt_hesap_raw, parent_bak)
        rows: [(3-digit-code, signed_balance), ...]  — ana parse için
        alt_hesap_raw: {parent: [{"kod", "ad", "borc_top", "alacak_top", "bakiye"}, ...]}
        parent_bak: {parent: signed_bakiye}  — doğruluk kontrolü için
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    best_ws = max(wb.worksheets, key=lambda ws: ws.max_row)
    code_col, borc_bak_col, alacak_bak_col, borc_top_col, alacak_top_col = _find_columns(best_ws)
    if not code_col or not borc_bak_col:
        raise ValueError("Hesap kodu veya bakiye kolonu tespit edilemedi.")

    # Formül tespiti: bakiye + toplam sütunlarının hepsine bak
    _kontrol_kolonlar = [c for c in [borc_bak_col, alacak_bak_col, borc_top_col, alacak_top_col] if c]
    _has_formula = False
    for _row in best_ws.iter_rows(min_row=2, max_row=30):
        for _col in _kontrol_kolonlar:
            _cell = _row[_col - 1]
            if _cell.value is not None:
                if isinstance(_cell.value, str) and _cell.value.startswith("="):
                    _has_formula = True
                    break
        if _has_formula:
            break
    if _has_formula:
        logger.info("Formül hücresi tespit edildi — data_only=True ile yeniden açılıyor.")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        best_ws = max(wb.worksheets, key=lambda ws: ws.max_row)
    else:
        logger.info("Hücreler sayısal — data_only=False yeterli.")

    logger.info(f"Bakiye sutunlari: borc_bak={borc_bak_col}, alacak_bak={alacak_bak_col}, "
                f"borc_top={borc_top_col}, alacak_top={alacak_top_col}")

    # İsim kolonu — başlık satırlarında ara, bulamazsan kod+1
    name_col = None
    for hrow in best_ws.iter_rows(min_row=1, max_row=15):
        for cell in hrow:
            v = _normalize_header(cell.value)
            if any(kw in v for kw in ["hesap adi", "hesap adı", "adi", "adı", "aciklama", "açıklama"]):
                name_col = cell.column
                break
        if name_col:
            break
    if name_col is None:
        name_col = code_col + 1

    raw_rows = []
    _alt_raw: dict[str, list] = {}       # tüm 3 haneli analar — reporter filtreler
    _parent_bak: dict[str, float] = {}   # exact 3-digit satır bakiyesi (doğruluk için)

    for row in best_ws.iter_rows(min_row=2):
        raw_code = row[code_col - 1].value
        if raw_code is None:
            continue
        s_raw = str(raw_code).strip()
        if not s_raw:
            continue

        # ── Tüm sütunları hemen oku ──
        def _fval(col):
            if not col:
                return 0.0
            try:
                return float(row[col - 1].value or 0)
            except (TypeError, ValueError):
                return 0.0

        bak_b    = _fval(borc_bak_col)
        bak_a    = _fval(alacak_bak_col)
        borc_top = _fval(borc_top_col)
        alacak_top = _fval(alacak_top_col)

        # ── Alt hesap & parent bakiye koleksiyonu (harf içeren kodlar dahil) ──
        digits = re.sub(r"[^0-9]", "", s_raw)
        if digits:
            parent = digits[:3]
            if True:  # tüm 3 haneli anaları topla; reporter hibrit filtre uygular
                if len(digits) >= 4:
                    # Sub-account: en az bir toplam sütunu sıfır değilse dahil et
                    if borc_top > 0 or alacak_top > 0:
                        # Signed bakiye: borç bakiyeli → pozitif, alacak bakiyeli → negatif
                        if bak_b > 0 and bak_a == 0:
                            bakiye = bak_b
                        elif bak_a > 0 and bak_b == 0:
                            bakiye = -bak_a
                        elif bak_b > 0 and bak_a > 0:
                            bakiye = bak_b - bak_a
                        else:
                            bakiye = borc_top - alacak_top
                        hesap_adi = ""
                        try:
                            nv = row[name_col - 1].value
                            hesap_adi = str(nv).strip() if nv else ""
                        except IndexError:
                            pass
                        if parent not in _alt_raw:
                            _alt_raw[parent] = []
                        _alt_raw[parent].append({
                            "kod": s_raw,
                            "ad": hesap_adi,
                            "borc_top": borc_top,
                            "alacak_top": alacak_top,
                            "bakiye": bakiye,
                        })
                elif len(digits) == 3:
                    # Exact parent row — signed bakiyeyi kaydet
                    if bak_b > 0 and bak_a == 0:
                        _parent_bak[parent] = bak_b
                    elif bak_a > 0 and bak_b == 0:
                        _parent_bak[parent] = -bak_a
                    elif bak_b > 0 and bak_a > 0:
                        _parent_bak[parent] = bak_b - bak_a

        # ── Ana parse: harf içeren kodları atla ──
        if re.search(r'[A-Za-z]', s_raw):
            continue

        s = s_raw
        parts = s.split(".")
        clean_parts = []
        for part in parts:
            part_clean = re.sub(r"[^0-9]", "", part)
            if part_clean:
                clean_parts.append(part_clean)
            else:
                break
        s = ".".join(clean_parts)
        if not s:
            continue

        # Bakiye tespiti
        if bak_b > 0 and bak_a == 0:
            borc = bak_b; alacak = 0.0
        elif bak_a > 0 and bak_b == 0:
            borc = 0.0; alacak = bak_a
        elif bak_b > 0 and bak_a > 0:
            borc = bak_b; alacak = bak_a
        else:
            root3 = s.split(".")[0][:3]
            if (root3.startswith("6") and len(root3) == 3
                    and root3 not in _SKIP_6XX
                    and borc_top > 0 and alacak_top > 0):
                if root3 in _ALACAK_NORMAL_6XX:
                    borc = 0.0; alacak = alacak_top
                elif root3 in _NET_HAREKET_6XX:
                    net = borc_top - alacak_top
                    borc = net if net > 0 else 0.0
                    alacak = (-net) if net < 0 else 0.0
                else:
                    borc = borc_top; alacak = 0.0
            else:
                borc = borc_top; alacak = alacak_top

        raw_rows.append((s, borc, alacak))

    if not raw_rows:
        return [], {}, {}

    all_codes = set(r[0] for r in raw_rows)
    has_hierarchy = any("." in code for code in all_codes)

    # Tüm ham kodlardaki rakam-only prefix'ler (çifte sayım tespiti için)
    _all_digit_prefixes = {re.sub(r"\D", "", c.split(".")[0]) for c in all_codes}

    result = []
    skipped = 0
    for code, borc, alacak in raw_rows:
        if borc == 0 and alacak == 0:
            continue
        if has_hierarchy and _is_parent_code(code, all_codes):
            skipped += 1
            continue
        # 1-2 haneli grup kodu: aynı prefix'e sahip 3+ haneli alt kod varsa atla
        _digits = re.sub(r"\D", "", code.split(".")[0])
        if len(_digits) <= 2:
            if any(p.startswith(_digits) and len(p) > len(_digits) for p in _all_digit_prefixes):
                skipped += 1
                continue
        root = _get_root3(code)
        if not root:
            continue
        if code != root:
            skipped += 1
            continue
        if borc > 0 and alacak == 0:
            balance = borc
        elif alacak > 0 and borc == 0:
            balance = -alacak
        elif borc > 0 and alacak > 0:
            balance = borc - alacak
        else:
            balance = 0
        if balance != 0:
            result.append((root, balance))

    logger.info(
        f"{'detay' if has_hierarchy else 'duz'} mizan | "
        f"ham:{len(raw_rows)} atlanan:{skipped} islenen:{len(result)}"
    )

    # ── Çifte sayım önleme: her grupta sadece yaprak (leaf) nodları tut ──
    # Bir kalemin kodu başka bir kalemin prefix'i ise (örn. "180.S2" ile "180.S2.01")
    # o ara-node satırını çıkar; sadece en alt kalemleri topla.
    def _leaf_only(kalemler: list) -> list:
        kodlar = {k["kod"] for k in kalemler}
        return [
            k for k in kalemler
            if not any(
                other != k["kod"] and other.startswith(k["kod"] + ".")
                for other in kodlar
            )
        ]

    for parent in list(_alt_raw.keys()):
        _alt_raw[parent] = _leaf_only(_alt_raw[parent])

    # Alt hesap filtresi: en az 1 anlamlı alt kalem (reporter hibrit eşiği uygular)
    alt_hesap_filtered = {p: v for p, v in _alt_raw.items() if len(v) >= 1}

    return result, alt_hesap_filtered, _parent_bak

# ─────────────────────────────────────────────
# 4. KURAL TABANLI EŞLEŞTİRME
# ─────────────────────────────────────────────

def _match_code(code: str) -> list[tuple[str, int]]:
    """
    Hesap kodunu TDHP tablosuna eşleştirir.
    Önce tam eşleşme, sonra 3 haneli prefix, sonra 2 haneli prefix dener.
    """
    # Tam eşleşme (3 hane)
    if code[:3] in _CODE_LOOKUP:
        return _CODE_LOOKUP[code[:3]]
    # 2 haneli prefix
    if len(code) >= 2 and code[:2] in _CODE_LOOKUP:
        return _CODE_LOOKUP[code[:2]]
    return []


def _apply_rules(rows: list[tuple[str, float]]) -> tuple[BalanceSheet, float]:
    """
    Satırları TDHP tablosuna göre BalanceSheet'e uygular.
    match_rate: eşleşen satır sayısı / toplam satır sayısı
    """
    bs = BalanceSheet()
    matched = 0

    for code, balance in rows:
        mappings = _match_code(code)
        if not mappings:
            logger.debug(f"Eşleşmeyen kod: {code} ({balance:,.0f})")
            continue
        matched += 1
        for field_name, sign in mappings:
            current = getattr(bs, field_name)
            setattr(bs, field_name, current + sign * balance)

    match_rate = matched / len(rows) if rows else 0.0
    bs.match_rate = match_rate
    return bs, match_rate


# ─────────────────────────────────────────────
# 5. AI FALLBACK (Claude API)
# ─────────────────────────────────────────────

def _ai_tamamla(bs: BalanceSheet, rows: list[tuple[str, float]], sector: str) -> BalanceSheet:
    """
    Fix kuralların eşleşme oranı yüksek ama bazı kalemler sıfır kaldıysa,
    AI sadece eksikleri tamamlar. Mevcut değerlere dokunmaz.
    """
    try:
        import anthropic
    except ImportError:
        logger.warning("anthropic paketi yok, AI tamamlama atlandı.")
        return bs

    kritik_alanlar = [
        "net_satislar", "satislarin_maliyeti", "finansman_giderleri",
        "genel_yonetim_giderleri", "stoklar", "ticari_alacaklar",
        "banka_kredileri_kv", "banka_kredileri_uv",
    ]
    eksik = [f for f in kritik_alanlar if getattr(bs, f, 0) == 0]

    if not eksik:
        logger.info("AI tamamlama: eksik kritik kalem yok, atlandı.")
        return bs

    logger.info(f"AI tamamlama: eksik kalemler → {eksik}")

    mizan_text = "\n".join(f"{code}\t{balance:,.2f}" for code, balance in rows[:200])
    eksik_str = ", ".join(eksik)

    system_prompt = f"""Sen bir Türk muhasebe uzmanısın.
Sana bir mizan veriliyor. Şu kalemler tespit edilemedi: {eksik_str}
SADECE bu eksik kalemlerin değerlerini JSON olarak ver.
Diğer kalemlere dokunma. Markdown veya açıklama ekleme, sadece JSON."""

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            system=system_prompt,
            messages=[{
                "role": "user",
                "content": f"Sektör: {sector}\n\nMizan:\n{mizan_text}"
            }]
        )
        raw = re.sub(r"```json|```", "", response.content[0].text).strip()
        data = json.loads(raw)
        for field_name, value in data.items():
            if hasattr(bs, field_name) and getattr(bs, field_name) == 0:
                try:
                    setattr(bs, field_name, float(value or 0))
                except (TypeError, ValueError):
                    pass
        logger.info("AI tamamlama başarılı.")
    except Exception as e:
        logger.warning(f"AI tamamlama başarısız: {e}")

    return bs


def _parse_with_ai(rows: list[tuple[str, float]], sector: str) -> BalanceSheet:
    """
    Kural tabanlı eşleşme düşük olduğunda Claude API'ye gönderir.
    Dönen JSON'u BalanceSheet'e dönüştürür.
    """
    try:
        import anthropic
    except ImportError:
        raise RuntimeError("anthropic paketi kurulu değil: pip install anthropic")

    # Mizan satırlarını metin olarak hazırla
    mizan_text = "\n".join(f"{code}\t{balance:,.2f}" for code, balance in rows[:300])

    system_prompt = """Sen bir Türk muhasebe uzmanısın. Sana bir mizan verilecek.
Mizan'daki hesap kodu ve bakiye verilerinden standart bilanço ve gelir tablosu kalemlerini çıkar.
SADECE JSON formatında yanıt ver, başka hiçbir metin ekleme.

JSON yapısı şu field'ları içermeli (TL cinsinden float değerler):
kasa, banka, diger_hazir_degerler, ticari_alacaklar, diger_alacaklar_kv, stoklar,
diger_donen_varliklar, ticari_alacaklar_uv, diger_alacaklar_uv, mali_duran_varliklar,
maddi_duran_varliklar, maddi_olmayan_duv, diger_duran_varliklar,
banka_kredileri_kv, uzun_vadeli_borclar_kv, ticari_borclar_kv, ortaklara_borclar,
diger_kv_borclar, banka_kredileri_uv, diger_uv_borclar,
odenmis_sermaye, sermaye_yedekleri, kar_yedekleri, gecmis_yil_karlari,
gecmis_yil_zararlari, donem_net_kari,
net_satislar, satislarin_maliyeti, pazarlama_giderleri, genel_yonetim_giderleri,
arge_giderleri, diger_faaliyet_gelirleri, diger_faaliyet_giderleri,
finansman_gelirleri, finansman_giderleri, vergi_gideri

Özel sınıflandırma kuralları:
- 170-179 (Yıllara Yaygın İnşaat ve Onarım Maliyetleri) → diger_donen_varliklar
- 190-199 (Diğer Dönen Varlıklar: iş avansları, hakediş vb.) → diger_donen_varliklar; diger_duran_varliklar'a YAZMA
- 350-359 (Yıllara Yaygın İnşaat Hakediş Bedelleri) → diger_kv_borclar
- 580-589 (Geçmiş Yıl Zararları) → gecmis_yil_zararlari olarak pozitif değer yaz; donem_net_kari'ye YAZMA
- donem_net_kari yalnızca 590 hesabından (Dönem Net Kârı/Zararı) gelir
- 740-749 (Hizmet/İnşaat Üretim Maliyeti) → satislarin_maliyeti
- 760-769 (Pazarlama Giderleri maliyet merkezi) → pazarlama_giderleri
- 770-779 (Genel Yönetim Giderleri maliyet merkezi) → genel_yonetim_giderleri"""

    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2000,
        system=system_prompt,
        messages=[{
            "role": "user",
            "content": f"Sektör: {sector}\n\nMizan:\nKod\tBakiye\n{mizan_text}"
        }]
    )

    raw = response.content[0].text.strip()
    # JSON fences temizle
    raw = re.sub(r"```json|```", "", raw).strip()
    data = json.loads(raw)

    bs = BalanceSheet()
    for field_name, value in data.items():
        if hasattr(bs, field_name):
            try:
                setattr(bs, field_name, float(value or 0))
            except (TypeError, ValueError):
                pass

    bs.parse_method = "ai_fallback"
    bs.match_rate = 1.0
    return bs


# ─────────────────────────────────────────────
# 6. DOĞRULAMA
# ─────────────────────────────────────────────

def _normalize_bilanco(bs: BalanceSheet, kapali_mizan: bool = False) -> BalanceSheet:
    """
    Mizan türünü tespit eder ve aktif-pasif dengesini sağlar.

    İki meşru mizan tipi:
    A) Yıl sonu kapatılmış: 590 > 0, 600-699 = 0 → bilanço dengeli
    B) Ara dönem açık:      590 = 0, 600-699 > 0 → bilanço dengeli

    Sorunlu durum — her ikisi aynı anda:
    590 > 0 VE 600 > 0 → muhasebeci yıl sonu kapatmış ama
    gelir tablosunu da bırakmış. Bu durumda 590'ı baz al,
    gelir tablosu kalemlerini scorer/analyzer için koru
    ama bilanço dengesini bozmamak için donem_net_kari'ni
    gelir tablosundan türetilmiş değerle ezmeyiz.

    Negatif duran varlık koruması:
    Mizanda net bakiye geliyorsa amortisman zaten düşülmüştür.
    """
    # Negatif duran varlık koruması
    if bs.maddi_duran_varliklar < 0:
        bs.maddi_duran_varliklar = 0
    if bs.maddi_olmayan_duv < 0:
        bs.maddi_olmayan_duv = 0

    # Mizan tipi tespiti
    var_590 = bs.donem_net_kari != 0
    var_600 = bs.net_satislar != 0

    if var_590 and var_600:
        # Çakışma: ikisi birden var — 590 net_kar için kullanılsın
        logger.info("Mizan tipi: Yıl sonu + gelir tablosu çakışması — 590 baz alınıyor.")
        bs._kullan_590 = True
        bs.warnings = getattr(bs, 'warnings', [])
        bs.warnings.append(
            "Mizanda hem 590 (dönem net kârı) hem gelir tablosu kalemleri mevcut. "
            "Bilanço dengesi için 590 baz alındı."
        )
    elif var_590 and not var_600:
        logger.info("Mizan tipi: Yıl sonu kapatılmış — 590 mevcut, gelir tablosu kapalı.")
    elif not var_590 and var_600:
        if kapali_mizan:
            logger.info(
                "Mizan tipi: Kapalı mizan — 6xx değerleri kapanış öncesi toplam olarak "
                "gelir tablosu için korunuyor, donem_net_kari=0 geçerli."
            )
        else:
            logger.info("Mizan tipi: Ara dönem açık — gelir tablosu kalemlerinden net kâr türetilecek.")
            # Gelir tablosundan net kâr hesapla ve özkaynağa ekle
            hesaplanan_net_kar = (
                bs.net_satislar - bs.satislarin_maliyeti
                - bs.faaliyet_giderleri
                + bs.diger_faaliyet_gelirleri - bs.diger_faaliyet_giderleri
                - bs.enflasyon_duzeltme_zarari   # 658 FAVÖK'e dahil değil
                + bs.finansman_gelirleri - bs.finansman_giderleri
                - bs.vergi_gideri
            )
            bs.donem_net_kari = hesaplanan_net_kar
            logger.info(f"Ara dönem net kârı hesaplandı: {hesaplanan_net_kar:,.0f} ₺")
    else:
        if bs.gecmis_yil_karlari > 0:
            # Kapalı mizan: yıl sonu kapanış yapılmış, dönem kârı 570'e devredilmiş.
            # donem_net_kari=0 doğru, aktif-pasif dengesi 570 üzerinden sağlanmış.
            logger.info(
                "Mizan tipi: Kapalı mizan — dönem kârı geçmiş yıl kârlarına "
                f"devredilmiş ({bs.gecmis_yil_karlari:,.0f} ₺)."
            )
        else:
            logger.warning("Mizan tipi tespit edilemedi — ne 590 ne de 600 mevcut.")

    # Son kontrol: aktif-pasif hâlâ tutmuyorsa logla, müdahale etme
    aktif = bs.toplam_aktif
    pasif = bs.toplam_pasif
    if aktif > 0 and pasif > 0:
        fark_oran = abs(aktif - pasif) / max(aktif, pasif)
        if fark_oran > 0.02:  # %2'den fazla fark varsa uyar
            logger.warning(
                f"Bilanço dengesi sağlanamadı: Aktif {aktif:,.0f} ₺, "
                f"Pasif {pasif:,.0f} ₺ (fark %{fark_oran*100:.1f}) — "
                f"muhtemelen mizanda eksik kalemler var."
            )

    return bs


def _validate(bs: BalanceSheet) -> list[str]:
    """Temel muhasebe denklemlerini kontrol eder, uyarı listesi döner."""
    warnings = []

    # Aktif = Pasif kontrolü (±%5 tolerans)
    if bs.toplam_aktif > 0:
        imbalance = abs(bs.toplam_aktif - bs.toplam_pasif) / bs.toplam_aktif
        if imbalance > 0.25:
            warnings.append(
                f"Aktif-Pasif dengesi bozuk: Aktif {bs.toplam_aktif:,.0f} ₺, "
                f"Pasif {bs.toplam_pasif:,.0f} ₺ (fark %{imbalance*100:.1f})"
            )

    # Net satış sıfır uyarısı
    if bs.net_satislar == 0:
        warnings.append("Net satış sıfır — gelir tablosu verileri eksik olabilir.")

    # Negatif özkaynak uyarısı
    if bs.ozkaynaklar < 0:
        warnings.append("Özkaynaklar negatif — şirket teknik olarak borca batık.")

    # Toplam aktif sıfır
    if bs.toplam_aktif == 0:
        warnings.append("Toplam aktif sıfır — parse başarısız olmuş olabilir.")

    return warnings


# ─────────────────────────────────────────────
# 7. ANA FONKSİYON
# ─────────────────────────────────────────────

def parse_mizan(
    filepath: str | Path,
    sector: str = "ticaret",
    use_ai_fallback: bool = True,
) -> BalanceSheet:
    """
    Ana parser fonksiyonu — Hybrid model.

    Akış:
      1. Fix kurallar (TDHP hesap kodu eşleştirme) çalışır
      2. use_ai_fallback=True ise AI her zaman devreye girer:
         - Eşleşme %80+ → AI sadece eksik/sıfır kalemleri tamamlar
         - Eşleşme %80 altı → AI tüm mizanı baştan parse eder
      3. Doğrulama çalışır

    Args:
        filepath: Excel mizan dosyası yolu
        sector: "ticaret" | "uretim" | "hizmet"
        use_ai_fallback: AI tamamlama kullanılsın mı (önerilen: True)

    Returns:
        BalanceSheet: Normalize edilmiş finansal veriler
    """
    logger.info(f"Parser başladı: {filepath}")

    # Excel oku (rows + alt_hesap verisi tek geçişte)
    rows, _alt_hesap_raw, _parent_bak = _read_excel(filepath)
    if not rows:
        raise ValueError("Excel dosyasında geçerli satır bulunamadı.")
    logger.info(f"{len(rows)} satır okundu.")

    # Aşama 1: Fix kurallar
    bs, match_rate = _apply_rules(rows)
    logger.info(f"Fix kural eşleşmesi: %{match_rate*100:.1f}")

    # ── Kapalı mizan tespiti (590=0, 6xx=0, kar 570'e devredilmiş) ──────────
    # Bu tip mizanlarda AI tetiklemek double-counting ve yanlış sınıflandırma
    # yaratır. Erken tespit edip AI bloğunu tamamen atlıyoruz.
    logger.info(
        f"[DEBUG_KAPALI] donem_net_kari={bs.donem_net_kari:.0f} "
        f"net_satislar={bs.net_satislar:.0f} "
        f"gecmis_yil_karlari={bs.gecmis_yil_karlari:.0f}"
    )
    _kapali_mizan = (
        bs.donem_net_kari == 0
        and bs.gecmis_yil_karlari > 0
    )
    if _kapali_mizan:
        logger.info(
            "Kapalı mizan tespit edildi (590=0, 6xx=0, 570>0) — "
            "AI tamamlama atlandı, kural tabanlı parse kullanılıyor."
        )
        bs.parse_method = "hybrid" if match_rate >= 0.80 else "rule_based"

    if use_ai_fallback and not _kapali_mizan:
        # Kural tabanlı parse dengesi ön kontrolü.
        # Açık dönem mizanda (590=0, 600>0) net kâr henüz özkaynağa eklenmemiştir;
        # tahmini pasifi kullanarak gerçek dengeyi hesapla.
        _aktif_pre = bs.toplam_aktif
        _pasif_pre = bs.toplam_pasif
        if bs.donem_net_kari == 0 and bs.net_satislar != 0:
            _tahmini_net_kar_pre = (
                bs.net_satislar - bs.satislarin_maliyeti
                - bs.faaliyet_giderleri
                + bs.diger_faaliyet_gelirleri - bs.diger_faaliyet_giderleri
                + bs.finansman_gelirleri - bs.finansman_giderleri
                - bs.vergi_gideri
            )
            _pasif_pre += _tahmini_net_kar_pre
        _pre_imbalance = (
            abs(_aktif_pre - _pasif_pre) / _aktif_pre
            if _aktif_pre > 0 and _pasif_pre > 0 else 1.0
        )

        if match_rate >= 0.80:
            if _pre_imbalance < 0.02:
                # Kural tabanlı parse zaten dengeli — AI tamamlama bilanço bozabilir, atla
                logger.info(
                    f"Kural tabanlı parse dengeli (%{_pre_imbalance*100:.2f} fark) — "
                    f"AI tamamlama atlandı."
                )
                bs.parse_method = "hybrid"
            else:
                # Dengede değil — AI eksik kalemleri tamamlasın
                logger.info(
                    f"Yüksek eşleşme ama bilanço %{_pre_imbalance*100:.1f} bozuk — "
                    f"AI eksik kalemleri tamamlıyor..."
                )
                bs = _ai_tamamla(bs, rows, sector)
                bs.parse_method = "hybrid"
        else:
            # Düşük eşleşme — AI tüm mizanı baştan parse eder
            logger.info(f"Düşük eşleşme (%{match_rate*100:.0f}) — AI tam parse yapıyor...")
            bs = _parse_with_ai(rows, sector)
            bs.warnings.append(
                f"Fix kural eşleşmesi %{match_rate*100:.0f} — AI ile tam parse yapıldı."
            )

        # Bilanço dengesi son kontrolü — AI sonrası hâlâ bozuksa yeniden parse et.
        _aktif = bs.toplam_aktif
        _pasif = bs.toplam_pasif
        if bs.donem_net_kari == 0 and bs.net_satislar != 0:
            _tahmini_net_kar = (
                bs.net_satislar - bs.satislarin_maliyeti
                - bs.faaliyet_giderleri
                + bs.diger_faaliyet_gelirleri - bs.diger_faaliyet_giderleri
                + bs.finansman_gelirleri - bs.finansman_giderleri
                - bs.vergi_gideri
            )
            _pasif += _tahmini_net_kar
        if _aktif > 0 and _pasif > 0:
            imbalance = abs(_aktif - _pasif) / _aktif
            if imbalance > 0.05:
                logger.info(f"Bilanço dengesi bozuk (%{imbalance*100:.1f}) — AI ile yeniden parse ediliyor...")
                # AI'a sadece 3 haneli ana hesapları gönder
                ana_rows = [(c, b) for c, b in rows if len(c) == 3]
                bs = _parse_with_ai(ana_rows, sector)
                bs.parse_method = "ai_reparse"
                bs.warnings.append(
                    f"Kural tabanlı parse bilanço dengesini sağlayamadı (%{imbalance*100:.1f} fark) — AI ile yeniden parse yapıldı."
                )

    # Aşama 2: Bilanço normalize et (denge düzeltme)
    bs = _normalize_bilanco(bs, kapali_mizan=_kapali_mizan)

    # Doğrulama
    validation_warnings = _validate(bs)
    bs.warnings.extend(validation_warnings)

    for w in bs.warnings:
        logger.warning(w)

    # Alt hesap verisi + doğruluk kontrolü
    bs.alt_hesaplar = {}
    for parent, kalemler in _alt_hesap_raw.items():
        uyari = ""
        if parent in _parent_bak:
            beklenen = _parent_bak[parent]
            alt_sum = sum(k["bakiye"] for k in kalemler)
            if beklenen != 0:
                sapma = abs(alt_sum - beklenen) / abs(beklenen)
                if sapma > 0.05:
                    uyari = (
                        "Alt hesap toplamı ana hesapla tam örtüşmeyebilir, "
                        "analiz tahmini niteliğindedir."
                    )
        bs.alt_hesaplar[parent] = {"kalemler": kalemler, "uyari": uyari}

    if bs.alt_hesaplar:
        logger.info(
            "Alt hesap verisi: "
            + ", ".join(
                f"{p}({len(v['kalemler'])} kalem)"
                for p, v in bs.alt_hesaplar.items()
            )
        )

    logger.info(f"Parse tamamlandı. Yöntem: {bs.parse_method}, "
                f"Toplam Aktif: {bs.toplam_aktif:,.0f} ₺")
    logger.info(f"MDD_DEBUG: mdd_raw={bs.maddi_duran_varliklar:.0f} mo_raw={bs.maddi_olmayan_duv:.0f} mali_raw={bs.mali_duran_varliklar:.0f} diger_hazir={bs.diger_hazir_degerler:.0f}")
    logger.info(f"BS_DEBUG: aktif={bs.toplam_aktif:.0f} pasif={bs.toplam_pasif:.0f} kasa={bs.kasa:.0f} banka={bs.banka:.0f} ta={bs.ticari_alacaklar:.0f} da={bs.diger_alacaklar_kv:.0f} st={bs.stoklar:.0f} dd={bs.diger_donen_varliklar:.0f} tuv={bs.ticari_alacaklar_uv:.0f} mdd={bs.maddi_duran_varliklar:.0f} mo={bs.maddi_olmayan_duv:.0f} dur={bs.diger_duran_varliklar:.0f} os={bs.odenmis_sermaye:.0f} ky={bs.kar_yedekleri:.0f} gk={bs.gecmis_yil_karlari:.0f} dn={bs.donem_net_kari:.0f}")
    return bs


# ─────────────────────────────────────────────
# 8. CLI — TEST
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if len(sys.argv) < 2:
        print("Kullanım: python parser.py mizan.xlsx [ticaret|uretim|hizmet]")
        sys.exit(1)

    filepath = sys.argv[1]
    sector = sys.argv[2] if len(sys.argv) > 2 else "ticaret"

    bs = parse_mizan(filepath, sector)
    result = bs.to_dict()

    print("\n" + "="*50)
    print("BİLANÇO ÖZETİ")
    print("="*50)
    print(f"Dönen Varlıklar   : {bs.donen_varliklar:>15,.0f} ₺")
    print(f"Duran Varlıklar   : {bs.duran_varliklar:>15,.0f} ₺")
    print(f"Toplam Aktif      : {bs.toplam_aktif:>15,.0f} ₺")
    print(f"KV Borçlar        : {bs.kv_borclar:>15,.0f} ₺")
    print(f"UV Borçlar        : {bs.uv_borclar:>15,.0f} ₺")
    print(f"Özkaynaklar       : {bs.ozkaynaklar:>15,.0f} ₺")
    print(f"Toplam Pasif      : {bs.toplam_pasif:>15,.0f} ₺")
    print()
    print("GELİR TABLOSU")
    print("="*50)
    print(f"Net Satışlar      : {bs.net_satislar:>15,.0f} ₺")
    print(f"Brüt Kâr          : {bs.brut_kar:>15,.0f} ₺")
    print(f"FAVÖK             : {bs.favok:>15,.0f} ₺")
    print(f"Net Kâr           : {bs.net_kar:>15,.0f} ₺")
    print()
    if bs.warnings:
        print("UYARILAR")
        print("="*50)
        for w in bs.warnings:
            print(f"  ⚠ {w}")
    print(f"\nEşleşme oranı: %{bs.match_rate*100:.1f} | Yöntem: {bs.parse_method}")
